"""XLSX to CSV/JSON conversion with hyperlink preservation."""

import csv
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import click

# Optional dependency: openpyxl for XLSX support
try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import Cell

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


def require_xlsx() -> None:
    """Raise error if openpyxl is not available."""
    if not XLSX_AVAILABLE:
        raise click.ClickException(
            "XLSX support requires openpyxl. Install with: pip install pipedrive-cli[xlsx]"
        )


@dataclass
class ConvertStats:
    """Statistics for conversion operation."""

    total_rows: int = 0
    total_columns: int = 0
    hyperlinks_found: int = 0
    hyperlinks_preserved: int = 0


@dataclass
class ConvertResult:
    """Result of a conversion operation."""

    records: list[dict[str, Any]] = field(default_factory=list)
    fieldnames: list[str] = field(default_factory=list)
    stats: ConvertStats = field(default_factory=ConvertStats)


def get_cell_value(cell: "Cell", preserve_links: bool = False) -> Any:
    """Get value from a cell, optionally preserving hyperlinks.

    Args:
        cell: openpyxl Cell object
        preserve_links: If True, return hyperlink URL instead of display value

    Returns:
        Cell value (URL if hyperlink and preserve_links, else display value)
    """
    if preserve_links and cell.hyperlink:
        # Return the hyperlink target URL
        return cell.hyperlink.target
    return cell.value


def load_xlsx(
    path: Path,
    sheet: str | None = None,
    header_row: int = 1,
    preserve_links: bool = False,
) -> ConvertResult:
    """Load data from an XLSX file.

    Args:
        path: Path to the XLSX file
        sheet: Sheet name (default: first sheet)
        header_row: Row number containing headers (1-indexed, default: 1)
        preserve_links: If True, replace cell values with hyperlink URLs

    Returns:
        ConvertResult with records, fieldnames, and stats
    """
    require_xlsx()

    wb = load_workbook(path, data_only=False)

    # Select sheet
    if sheet:
        if sheet not in wb.sheetnames:
            raise click.ClickException(
                f"Sheet '{sheet}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            )
        ws = wb[sheet]
    else:
        ws = wb.active

    stats = ConvertStats()

    # Get headers from specified row
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    fieldnames = [str(cell.value or f"column_{i}") for i, cell in enumerate(header_cells, 1)]
    stats.total_columns = len(fieldnames)

    # Read data rows
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1):
        # Skip empty rows
        if all(cell.value is None for cell in row):
            continue

        record: dict[str, Any] = {}
        for i, cell in enumerate(row):
            if i >= len(fieldnames):
                break

            field_name = fieldnames[i]

            # Track hyperlinks
            if cell.hyperlink:
                stats.hyperlinks_found += 1
                if preserve_links:
                    stats.hyperlinks_preserved += 1

            value = get_cell_value(cell, preserve_links)

            # Convert to appropriate type
            if value is None:
                record[field_name] = ""
            elif isinstance(value, (int, float)):
                record[field_name] = value
            else:
                record[field_name] = str(value)

        records.append(record)

    stats.total_rows = len(records)

    return ConvertResult(records=records, fieldnames=fieldnames, stats=stats)


def write_csv(records: list[dict[str, Any]], fieldnames: list[str], output: Path) -> None:
    """Write records to a CSV file.

    Args:
        records: List of record dicts
        fieldnames: Column names in order
        output: Output file path
    """
    with open(output, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            # Flatten complex values to JSON strings
            flat_record: dict[str, Any] = {}
            for key, value in record.items():
                if isinstance(value, (dict, list)):
                    flat_record[key] = json.dumps(value, ensure_ascii=False)
                else:
                    flat_record[key] = value
            writer.writerow(flat_record)


def write_json(records: list[dict[str, Any]], output: Path) -> None:
    """Write records to a JSON file.

    Args:
        records: List of record dicts
        output: Output file path
    """
    with open(output, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)


def detect_output_format(output: Path) -> str:
    """Detect output format from file extension.

    Args:
        output: Output file path

    Returns:
        Format string: "csv" or "json"

    Raises:
        ClickException if format cannot be determined
    """
    suffix = output.suffix.lower()
    if suffix == ".csv":
        return "csv"
    elif suffix == ".json":
        return "json"
    else:
        raise click.ClickException(
            f"Cannot determine output format from extension '{suffix}'. "
            "Use --format to specify 'csv' or 'json'."
        )
